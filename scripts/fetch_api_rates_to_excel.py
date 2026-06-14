from __future__ import annotations

import argparse
import json
import os
from datetime import UTC, date, datetime
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT_DIR / "data" / "currency_rates.xlsx"
DATA_SHEET = "currency_rates"
HEADERS = [
    "date",
    "currency",
    "currency_name",
    "buy_rate",
    "sell_rate",
    "source",
    "comment",
]
CURRENCY_NAMES = {
    "USD": "US Dollar",
    "EUR": "Euro",
    "PLN": "Polish Zloty",
}
API_SOURCES = {"НБУ API", "Minfin Mizhbank API"}


def request_json(url: str) -> object:
    request = Request(url, headers={"User-Agent": "currency-rates-github-pages/1.0"})

    with urlopen(request, timeout=30) as response:
        return json.loads(response.read().decode("utf-8"))


def fetch_nbu_rate(rate_date: date, currency: str) -> dict[str, str | float] | None:
    query = urlencode(
        {
            "valcode": currency,
            "date": rate_date.strftime("%Y%m%d"),
            "json": "",
        }
    )
    url = f"https://bank.gov.ua/NBUStatService/v1/statdirectory/exchange?{query}"
    payload = request_json(url)

    if not isinstance(payload, list) or not payload:
        return None

    item = payload[0]
    rate = float(item["rate"])

    return {
        "date": rate_date.isoformat(),
        "currency": currency,
        "currency_name": CURRENCY_NAMES.get(currency, str(item.get("txt", currency))),
        "buy_rate": rate,
        "sell_rate": rate,
        "source": "НБУ API",
        "comment": f"Official NBU exchange rate fetched at {datetime.now(UTC).date().isoformat()}",
    }


def fetch_minfin_mizhbank_rate(
    rate_date: date,
    currency: str,
    api_key: str,
) -> dict[str, str | float] | None:
    url = (
        f"https://api.minfin.com.ua/mb/{api_key}/{rate_date.isoformat()}/"
        f"?{urlencode({'currency': currency.lower()})}"
    )
    payload = request_json(url)

    if isinstance(payload, dict):
        items = [payload]
    elif isinstance(payload, list):
        items = payload
    else:
        items = []

    if not items:
        return None

    item = items[0]
    buy_rate = float(item["bid"])
    sell_rate = float(item["ask"])
    point_date = item.get("pointDate") or item.get("date") or rate_date.isoformat()

    return {
        "date": str(point_date)[:10],
        "currency": currency,
        "currency_name": CURRENCY_NAMES.get(currency, currency),
        "buy_rate": buy_rate,
        "sell_rate": sell_rate,
        "source": "Minfin Mizhbank API",
        "comment": f"Minfin interbank rate for {currency}; source timestamp: {point_date}",
    }


def remove_existing_api_rows(sheet) -> None:
    for row_index in range(sheet.max_row, 1, -1):
        source = sheet.cell(row_index, 6).value

        if source in API_SOURCES:
            sheet.delete_rows(row_index)


def append_records(records: list[dict[str, str | float]]) -> None:
    workbook = load_workbook(EXCEL_PATH)

    if DATA_SHEET not in workbook.sheetnames:
        raise ValueError(f"Missing required sheet: {DATA_SHEET}")

    sheet = workbook[DATA_SHEET]
    headers = [cell.value for cell in sheet[1]][: len(HEADERS)]

    if headers != HEADERS:
        raise ValueError(f"Expected columns {HEADERS}, found {headers}")

    remove_existing_api_rows(sheet)

    for record in records:
        sheet.append([record[column] for column in HEADERS])

    workbook.save(EXCEL_PATH)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch NBU and Minfin API rates into Excel.")
    parser.add_argument(
        "--date",
        default=date.today().isoformat(),
        help="Rate date in YYYY-MM-DD format. Defaults to today.",
    )
    parser.add_argument(
        "--currencies",
        nargs="+",
        default=["USD", "EUR"],
        help="Currency codes to fetch. Defaults to USD EUR.",
    )
    parser.add_argument(
        "--sources",
        nargs="+",
        choices=["nbu", "minfin"],
        default=["nbu", "minfin"],
        help="API sources to fetch.",
    )
    return parser.parse_args()


def main() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Missing Excel file: {EXCEL_PATH}")

    args = parse_args()
    rate_date = datetime.strptime(args.date, "%Y-%m-%d").date()
    currencies = [currency.upper() for currency in args.currencies]
    minfin_api_key = os.environ.get("MINFIN_API_KEY")
    records = []

    if "minfin" in args.sources and not minfin_api_key:
        raise RuntimeError("Set MINFIN_API_KEY before fetching Minfin Mizhbank API rates.")

    for currency in currencies:
        if "nbu" in args.sources:
            nbu_record = fetch_nbu_rate(rate_date, currency)

            if nbu_record:
                records.append(nbu_record)

        if "minfin" in args.sources:
            minfin_record = fetch_minfin_mizhbank_rate(rate_date, currency, minfin_api_key)

            if minfin_record:
                records.append(minfin_record)

    append_records(records)
    print(f"Added {len(records)} API records to {EXCEL_PATH}")


if __name__ == "__main__":
    main()
