from __future__ import annotations

import json
from datetime import UTC, date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT_DIR / "data" / "currency_rates.xlsx"
JSON_PATH = ROOT_DIR / "data" / "currency_rates.json"
CURRENCY_NAMES = {
    "USD": "US Dollar",
    "EUR": "Euro",
    "EURO": "Euro",
}
COMMENT = "Internal company calculated rate. Buy and sell are equal because the source table contains one rate."


def format_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    return str(value).strip()


def normalize_currency(value: object) -> str:
    code = str(value or "").strip().upper()
    return "EUR" if code == "EURO" else code


def parse_rate(value: object) -> float | None:
    if value is None or value == "":
        return None

    return float(value)


def build_record(rate_date: object, currency_code: object, rate: float, source: str) -> dict[str, str | float]:
    normalized_code = normalize_currency(currency_code)

    return {
        "date": format_date(rate_date),
        "currency": normalized_code,
        "currency_name": CURRENCY_NAMES.get(normalized_code, normalized_code),
        "buy_rate": rate,
        "sell_rate": rate,
        "source": source,
        "comment": COMMENT,
    }


def main() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Missing Excel file: {EXCEL_PATH}")

    workbook = load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    source = str(sheet["B1"].value or "Excel").strip()
    currency_columns = [
        (sheet["B2"].value, 2),
        (sheet["C2"].value, 3),
    ]

    records = []

    for row in sheet.iter_rows(min_row=3, max_col=3, values_only=True):
        rate_date = row[0]

        if not rate_date:
            continue

        for currency_code, column_index in currency_columns:
            if not currency_code:
                continue

            rate = parse_rate(row[column_index - 1])

            if rate is None:
                continue

            records.append(build_record(rate_date, currency_code, rate, source))

    payload = {
        "source": "data/currency_rates.xlsx",
        "sheet": sheet.title,
        "generated_at": datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "records": records,
    }

    JSON_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Converted {len(records)} records from first sheet '{sheet.title}' to {JSON_PATH}")


if __name__ == "__main__":
    main()
