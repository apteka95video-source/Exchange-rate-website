from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT_DIR / "data" / "currency_rates.xlsx"
OUTPUT_SHEET = "currency_rates"


def main() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Missing Excel file: {EXCEL_PATH}")

    workbook = load_workbook(EXCEL_PATH)
    source_sheet = workbook.worksheets[0]

    if OUTPUT_SHEET in workbook.sheetnames and workbook[OUTPUT_SHEET] is not source_sheet:
        del workbook[OUTPUT_SHEET]
        workbook.save(EXCEL_PATH)
        print(f"Removed old helper sheet '{OUTPUT_SHEET}'. Source sheet is '{source_sheet.title}'.")
        return

    print(f"No helper sheet to remove. Source sheet is '{source_sheet.title}'.")


if __name__ == "__main__":
    main()
